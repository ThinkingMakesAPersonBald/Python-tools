'''
Author: xinhua.pei xinhua.pei@airudder.com
Date: 2023-06-03 00:33:54
LastEditors: xinhua.pei xinhua.pei@airudder.com
LastEditTime: 2023-06-03 11:20:06
FilePath: /Python-tools/SyllableDivisionTool.py
Description: 这是默认设置,请设置`customMade`, 打开koroFileHeader查看配置 进行设置: https://github.com/OBKoro1/koro1FileHeader/wiki/%E9%85%8D%E7%BD%AE
'''
# success one
# import openpyxl
# import pyphen
# import os

# root_path = '/Users/peixinhua/Desktop/Junior High School Vocabulary'
# filename = 'Junior High School Vovabulary (Final Version).xlsx'

# # 读取 Excel 文件
# workbook = openpyxl.load_workbook(os.path.join(root_path, filename))
# worksheet = workbook.active

# # 创建 Pyphen 对象，用于将单词划分为音节
# dic = pyphen.Pyphen(lang='en')

# # 遍历每个单元格并将单词划分为音节，并将结果写入新的列中
# for row in worksheet.iter_rows():
#     for cell in row:
#         word = cell.value
#         if word is not None and isinstance(word, str):
#             # 将单词划分为音节
#             syllables = dic.inserted(word).split("-")
#             result = "-".join(syllables)
#             # 将音节连接起来并写入新的列中
#             new_cell = cell.offset(column=1)
#             new_cell.value = result

# # 将结果保存回 Excel 文件
# workbook.save(os.path.join(root_path, 'output.xlsx'))


import openpyxl
import os
import nltk

root_path = '/Users/peixinhua/Desktop/Junior High School Vocabulary'
filename = 'Junior High School Vovabulary (Final Version).xlsx'

# 读取 Excel 文件
workbook = openpyxl.load_workbook(os.path.join(root_path, filename))
worksheet = workbook.active

# 创建 SyllableTokenizer 对象，用于将单词划分为音节
tokenizer = nltk.tokenize.SyllableTokenizer()

# 遍历每个单元格并将单词划分为音节，并将结果写入新的列中
for row in worksheet.iter_rows():
    for cell in row:
        word = cell.value
        if word is not None and isinstance(word, str):
            # 将单词划分为音节
            syllables = tokenizer.tokenize(word)
            result = "-".join(syllables)
            # 将音节连接起来并写入新的列中
            new_cell = cell.offset(column=1)
            new_cell.value = result

# 将结果保存回 Excel 文件
workbook.save(os.path.join(root_path, 'output.xlsx'))
#结果需要另外处理一下
#1. '---' replace to '-'
#2. ' -' replace to '-'
#3. 't-h' replace to th
#4. '- ' replace to 
#5. '-'' replace to '''

# import openpyxl
# import os
# import nltk

# root_path = '/Users/peixinhua/Desktop/Junior High School Vocabulary'
# filename = 'Junior High School Vovabulary (Final Version).xlsx'

# # 读取 Excel 文件
# workbook = openpyxl.load_workbook(os.path.join(root_path, filename))
# worksheet = workbook.active

# # 加载 CMU Pronouncing Dictionary
# nltk.download('cmudict')
# cmudict = nltk.corpus.cmudict.dict()

# # 遍历每个单元格并将单词划分为音节，并将结果写入新的列中
# for row in worksheet.iter_rows():
#     for cell in row:
#         word = cell.value
#         if word is not None and isinstance(word, str):
#             # 将单词划分为音节
#             syllables = cmudict.get(word.lower(), [[]])[0]
#             result = "-".join(syllables)
#             # 将音节连接起来并写入新的列中
#             new_cell = cell.offset(column=1)
#             new_cell.value = result

# # 将结果保存回 Excel 文件
# workbook.save(os.path.join(root_path, 'output.xlsx'))
            