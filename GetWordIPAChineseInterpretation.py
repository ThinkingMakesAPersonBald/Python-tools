#获取英语 IPA音标及简单中文释义
import openpyxl
from openpyxl.utils import get_column_letter
import os

root_path = '/Users/peixinhua/Downloads'
file_name = 'example.xlsx'

# 加载Excel文件
wb = openpyxl.load_workbook(os.path.join(root_path,file_name))
sheet = wb['Sheet1']

# 初始化Google翻译API
from googletrans import Translator
translator = Translator()

# 获取单词的IPA音标和释义
for row in range(2, sheet.max_row + 1):
    word_cell = sheet['A' + str(row)]
    meaning_cell = sheet['B' + str(row)]
    if word_cell.value:
        # 获取单词的IPA音标
        word = word_cell.value
        ipa = translator.translate(word, dest='en').pronunciation
        sheet['C' + str(row)] = ipa
        
        # 获取单词的释义
        meaning = translator.translate(word, dest='zh-CN').text
        sheet['D' + str(row)] = meaning

# 保存Excel文件
wb.save(os.path.join(root_path,'example_with_ipa.xlsx'))